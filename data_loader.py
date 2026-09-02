"""
Data loader for Multibrand Daily Flash Reports.
Parses every Excel file in the folder and returns structured DataFrames.

Files may be added in any order. The report date is read from cell A2 inside
each workbook, never from the filename, and the combined result is sorted by
date at the end -- so dropping a June file in after an August one is fine.

When two files carry the same report date, the later download wins. "Later" is
decided by the browser's "[n]" filename suffix (which increments on every
re-download) and only falls back to modification time. That ordering matters:
Streamlit Cloud serves the app from a fresh git clone, which stamps every file
with the same checkout time, so mtime alone cannot rank same-date files there.

Duplicate dates whose contents actually disagree are recorded in LAST_LOAD
rather than silently resolved, so the dashboard can surface them.
"""

import os
import re
import glob
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

SHEET_NAME = 'Multibrand_DailyFlashReport'

# Bound the single pass over each sheet. Sheets seen so far end by row 87;
# the margin covers a layout that adds rows.
MAX_ROW = 120
MAX_COL = 21

# A store row in column A looks like "9501 - Normal".
_STORE_RE = re.compile(r'^\d{4}\s*-\s*')

# Each per-store block is identified by text in the header rows just above it.
_SECTION_KEYS = {
    'day sales': 'sales',
    'day trans': 'transactions',
    'average check': 'channels',
    'labor': 'labor',
}

# Populated by load_all_reports() so the UI can report on the load without
# re-reading anything. Keys: files_seen, files_parsed, files_unreadable,
# dates_loaded, duplicate_dates, conflicting_dates, errors.
LAST_LOAD = {}


def parse_date_from_cell(value):
    """Extract date from cell like 'Selected Date:2/22/2026'."""
    if not value:
        return None
    match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', str(value))
    if match:
        return datetime.strptime(match.group(1), '%m/%d/%Y').date()
    return None


def safe_float(val, default=0.0):
    """Safely convert value to float."""
    if val is None or val == '' or val == '-':
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    """Safely convert value to int."""
    if val is None or val == '' or val == '-':
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def download_sequence(filepath):
    """Order a file by when it was downloaded.

    Browsers suffix repeat downloads of the same name with '[1]', '[2]', ... so
    a higher number is a later pull of that report. The un-suffixed file is the
    first download and sorts earliest. This survives a git clone, which mtime
    does not.
    """
    match = re.search(r'\[(\d+)\]\.xlsx$', os.path.basename(filepath))
    return int(match.group(1)) if match else 0


def _read_grid(filepath):
    """Read the flash report sheet into a padded 2-D list in one pass.

    read_only mode streams the sheet instead of building a full cell graph,
    which is what makes parsing ~180 workbooks practical.
    """
    wb = load_workbook(filepath, data_only=True, read_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            return None
        ws = wb[SHEET_NAME]
        grid = [
            [c.value for c in row]
            for row in ws.iter_rows(min_row=1, max_row=MAX_ROW, max_col=MAX_COL)
        ]
    finally:
        wb.close()

    # A short or narrow sheet yields fewer rows/cells; pad so indexing is safe.
    for row in grid:
        if len(row) < MAX_COL:
            row.extend([None] * (MAX_COL - len(row)))
    while len(grid) < MAX_ROW:
        grid.append([None] * MAX_COL)
    return grid


def _cell(grid, row, col):
    """1-indexed cell access matching the original openpyxl ws.cell(row, col)."""
    if 1 <= row <= len(grid) and 1 <= col <= MAX_COL:
        return grid[row - 1][col - 1]
    return None


def find_sections(grid):
    """Locate each per-store block by the header text above it.

    The report ships in at least two layouts whose section rows differ by one.
    Hardcoded row ranges silently dropped the first store (9501 - Normal) from
    transactions, channel mix and labor on every file in the newer layout, so
    the blocks are anchored on their headers instead.
    """
    runs, current = [], []
    for row in range(1, len(grid) + 1):
        value = _cell(grid, row, 1)
        if value is not None and _STORE_RE.match(str(value).strip()):
            current.append(row)
        elif current:
            runs.append(current)
            current = []
    if current:
        runs.append(current)

    found = {}
    for run in runs:
        label = None
        for row in range(run[0] - 1, max(0, run[0] - 6), -1):
            joined = ' '.join(
                str(_cell(grid, row, col) or '') for col in range(1, MAX_COL + 1)
            ).lower()
            for key, name in _SECTION_KEYS.items():
                if key in joined:
                    label = name
                    break
            if label:
                break
        if label and label not in found:
            found[label] = run
    return found


def find_labeled_row(grid, text):
    """Row number whose column A starts with `text`, or None."""
    for row in range(1, len(grid) + 1):
        value = _cell(grid, row, 1)
        if value is not None and str(value).strip().lower().startswith(text.lower()):
            return row
    return None


def parse_flash_report(filepath):
    """
    Parse a single Multibrand Flash Report Excel file.
    Returns a dict with record lists for each data section, or None if the
    file carries no readable report date.
    """
    grid = _read_grid(filepath)
    if grid is None:
        return None

    report_date = parse_date_from_cell(_cell(grid, 2, 1))
    if not report_date:
        return None

    def f(row, col):
        return safe_float(_cell(grid, row, col))

    def i(row, col):
        return safe_int(_cell(grid, row, col))

    sections = find_sections(grid)

    # Sales / transactions share the same 21-column shape: five period blocks
    # (day, wtd, ptd, ytd, r13) of ty / ly / diff / pct.
    SPANS = [('day', 2), ('wtd', 6), ('ptd', 10), ('ytd', 14), ('r13', 18)]

    def block(row, prefix, conv):
        out = {}
        for name, start in SPANS:
            out[f'{name}_{prefix}_ty'] = conv(row, start)
            out[f'{name}_{prefix}_ly'] = conv(row, start + 1)
            out[f'{name}_{prefix}_diff'] = conv(row, start + 2)
            out[f'{name}_{prefix}_pct'] = f(row, start + 3)
        return out

    # --- Section 1: Sales by Store ---
    sales_records = []
    for row_num in sections.get('sales', []):
        store_name = _cell(grid, row_num, 1)
        if not store_name or 'Totals' in str(store_name) or 'Brand' in str(store_name):
            continue
        rec = {'date': report_date, 'store': str(store_name).strip()}
        rec.update(block(row_num, 'sales', f))
        sales_records.append(rec)

    # Brand totals — located by label, not a fixed row.
    totals_row = find_labeled_row(grid, 'Brand Totals') or find_labeled_row(grid, 'Totals')
    brand_total = {'date': report_date}
    if totals_row:
        brand_total.update(block(totals_row, 'sales', f))
    else:
        brand_total.update(block(0, 'sales', f))

    # --- Section 2: Transactions by Store ---
    trans_records = []
    for row_num in sections.get('transactions', []):
        store_name = _cell(grid, row_num, 1)
        if not store_name or 'Totals' in str(store_name):
            continue
        rec = {'date': report_date, 'store': str(store_name).strip()}
        rec.update(block(row_num, 'trans', i))
        trans_records.append(rec)

    # --- Section 3: Channel Mix ---
    channel_records = []
    for row_num in sections.get('channels', []):
        store_name = _cell(grid, row_num, 1)
        if not store_name or 'Totals' in str(store_name):
            continue
        channel_records.append({
            'date': report_date,
            'store': str(store_name).strip(),
            'avg_check_ty': f(row_num, 2),
            'avg_check_ly': f(row_num, 3),
            'avg_check_diff': f(row_num, 4),
            'avg_check_pct': f(row_num, 5),
            'dine_in_sales': f(row_num, 6),
            'dine_in_trans': i(row_num, 7),
            'dine_in_avg_check': f(row_num, 8),
            'dine_in_pct_sales': f(row_num, 9),
            'carry_out_sales': f(row_num, 10),
            'carry_out_trans': i(row_num, 11),
            'carry_out_avg_check': f(row_num, 12),
            'carry_out_pct_sales': f(row_num, 13),
            'delivery_sales': f(row_num, 14),
            'delivery_trans': i(row_num, 15),
            'delivery_avg_check': f(row_num, 16),
            'delivery_pct_sales': f(row_num, 17),
            'drive_thru_sales': f(row_num, 18),
            'drive_thru_trans': i(row_num, 19),
            'drive_thru_avg_check': f(row_num, 20),
            'drive_thru_pct_sales': f(row_num, 21),
        })

    # --- Section 4: Labor & 3rd Party ---
    labor_records = []
    for row_num in sections.get('labor', []):
        store_name = _cell(grid, row_num, 1)
        if not store_name or 'Totals' in str(store_name):
            continue
        labor_records.append({
            'date': report_date,
            'store': str(store_name).strip(),
            'labor_dollars': f(row_num, 2),
            'labor_pct': f(row_num, 3),
            'olo_sales': f(row_num, 4),
            'doordash': f(row_num, 5),
            'ubereats': f(row_num, 6),
            'grubhub': f(row_num, 7),
            'eatstreet': f(row_num, 8),
            'ezcater': f(row_num, 9),
            'total_3rd_party_dollars': f(row_num, 10),
            'total_3rd_party_pct': f(row_num, 11),
        })

    return {
        'date': report_date,
        'sales': sales_records,
        'brand_total': brand_total,
        'transactions': trans_records,
        'channels': channel_records,
        'labor': labor_records,
    }


def _payload_signature(result):
    """A comparable fingerprint of a parsed report, ignoring which file it came from.

    Floats are rounded before comparison: the same figure re-exported can differ
    in its last binary digit (28.242216290000002 vs 28.24221629), which is not a
    disagreement worth reporting.
    """
    def norm(value):
        if isinstance(value, float):
            return round(value, 6)
        if isinstance(value, dict):
            return {k: norm(v) for k, v in value.items()}
        if isinstance(value, list):
            return [norm(v) for v in value]
        return value

    return repr(norm([
        result['sales'], result['brand_total'], result['transactions'],
        result['channels'], result['labor'],
    ]))


def load_all_reports(folder_path=None):
    """
    Load every Multibrand Flash Report in the folder into DataFrames.

    Files are read in download order, so when several files carry the same
    report date the last one downloaded wins -- deterministically, and the same
    way locally as on Streamlit Cloud.
    """
    if folder_path is None:
        folder_path = os.path.dirname(os.path.abspath(__file__))

    pattern = os.path.join(folder_path, 'Multibrand_FlashReport*.xlsx')

    # Sort so that, for any given date, the newest download is parsed last and
    # therefore wins. mtime only breaks ties between equal sequence numbers.
    def order(path):
        try:
            mtime = os.path.getmtime(path)
        except OSError:
            mtime = 0.0
        return (download_sequence(path), mtime, os.path.basename(path))

    files = sorted(glob.glob(pattern), key=order)

    winners = {}       # date -> (result, filepath)
    duplicates = {}    # date -> [filenames], oldest first
    conflicts = {}     # date -> [filenames] whose contents disagree
    errors = []

    for filepath in files:
        name = os.path.basename(filepath)
        try:
            result = parse_flash_report(filepath)
        except Exception as exc:
            errors.append(f'{name}: {exc}')
            continue
        if result is None:
            errors.append(f'{name}: no readable report date in cell A2')
            continue

        rdate = result['date']
        if rdate in winners:
            duplicates.setdefault(rdate, [winners[rdate][1]]).append(name)
            if _payload_signature(result) != _payload_signature(winners[rdate][0]):
                conflicts.setdefault(rdate, [winners[rdate][1]]).append(name)
        winners[rdate] = (result, name)

    all_sales, all_brand, all_trans, all_chan, all_labor = [], [], [], [], []
    for rdate in sorted(winners):
        result = winners[rdate][0]
        all_sales.extend(result['sales'])
        all_brand.append(result['brand_total'])
        all_trans.extend(result['transactions'])
        all_chan.extend(result['channels'])
        all_labor.extend(result['labor'])

    data = {
        'sales': pd.DataFrame(all_sales),
        'brand_totals': pd.DataFrame(all_brand),
        'transactions': pd.DataFrame(all_trans),
        'channels': pd.DataFrame(all_chan),
        'labor': pd.DataFrame(all_labor),
    }
    for key in data:
        if not data[key].empty and 'date' in data[key].columns:
            data[key] = data[key].sort_values('date').reset_index(drop=True)

    LAST_LOAD.clear()
    LAST_LOAD.update({
        'folder': folder_path,
        'files_seen': len(files),
        'files_parsed': len(files) - len(errors),
        'dates_loaded': len(winners),
        'duplicate_dates': {d.isoformat(): n for d, n in sorted(duplicates.items())},
        'conflicting_dates': {d.isoformat(): n for d, n in sorted(conflicts.items())},
        'errors': errors,
    })

    if not files:
        print(f'No matching Excel files found in: {folder_path}')
    for message in errors:
        print(f'Error parsing {message}')
    for rdate, names in sorted(conflicts.items()):
        print(f'Conflicting reports for {rdate}: {names} -- using {names[-1]}')

    return data


if __name__ == '__main__':
    import json
    import time

    started = time.time()
    data = load_all_reports()
    elapsed = time.time() - started

    for key, df in data.items():
        print(f'\n{key}: {len(df)} rows')
        if not df.empty:
            print(df.head())

    print('\n--- load summary ---')
    print(json.dumps(LAST_LOAD, indent=2, default=str))
    print(f'elapsed: {elapsed:.1f}s')
